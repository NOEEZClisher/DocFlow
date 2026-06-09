from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from docflow.converters.registry import default_registry
from docflow.converters.xlsx_converter import XlsxConverter


def save_workbook(path: Path, workbook: Workbook) -> Path:
    workbook.save(path)
    return path


def test_xlsx_converter_converts_single_sheet_table(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "점수"
    sheet.append(["이름", "점수"])
    sheet.append(["민수", 95])
    path = save_workbook(tmp_path / "scores.xlsx", workbook)

    document = XlsxConverter().convert(path)

    assert document.kind == "xlsx"
    assert document.raw_text == "\n".join(
        [
            "## 점수",
            "",
            "| 이름 | 점수 |",
            "| --- | --- |",
            "| 민수 | 95 |",
        ]
    )
    assert document.rendered_html is not None
    assert "<h2>점수</h2>" in document.rendered_html
    assert "<table>" in document.rendered_html


def test_xlsx_converter_converts_multiple_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "1반"
    first.append(["이름", "상태"])
    first.append(["민수", "제출"])
    second = workbook.create_sheet("2반")
    second.append(["이름", "상태"])
    second.append(["지아", "미제출"])
    path = save_workbook(tmp_path / "classes.xlsx", workbook)

    document = XlsxConverter().convert(path)

    assert "## 1반" in document.raw_text
    assert "| 민수 | 제출 |" in document.raw_text
    assert "## 2반" in document.raw_text
    assert "| 지아 | 미제출 |" in document.raw_text


def test_xlsx_converter_handles_empty_cells_and_escapes_table_text(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "메모"
    sheet["A1"] = "항목"
    sheet["B1"] = "내용"
    sheet["A2"] = "초안"
    sheet["C2"] = "줄1\n줄2 | 확인"
    path = save_workbook(tmp_path / "notes.xlsx", workbook)

    document = XlsxConverter().convert(path)

    assert "| 항목 | 내용 |  |" in document.raw_text
    assert "| 초안 |  | 줄1 줄2 \\| 확인 |" in document.raw_text


def test_xlsx_converter_marks_empty_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "빈문서"
    path = save_workbook(tmp_path / "empty.xlsx", workbook)

    document = XlsxConverter().convert(path)

    assert document.raw_text == "## 빈문서\n\n빈 시트"


def test_registry_selects_xlsx_converter_for_xlsx_files() -> None:
    converter = default_registry.get_converter(Path("scores.xlsx"))

    assert isinstance(converter, XlsxConverter)
