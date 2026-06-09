from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from docflow.core.document_model import Document, DocumentLoadError
from docflow.core.markdown_renderer import render_markdown

from .base import DocumentConverter


class XlsxConverter(DocumentConverter):
    """Convert XLSX worksheets into Markdown sections with tables."""

    supported_extensions = frozenset({".xlsx"})

    def convert(self, path: Path) -> Document:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise DocumentLoadError("openpyxl 패키지가 설치되어 있지 않습니다.") from exc

        try:
            workbook = load_workbook(path, data_only=False, read_only=True)
        except Exception as exc:
            raise DocumentLoadError(f"XLSX 파일을 읽을 수 없습니다: {path}") from exc

        try:
            raw_text = "\n\n".join(_worksheet_to_markdown(sheet) for sheet in workbook.worksheets).strip()
        finally:
            workbook.close()

        return Document(
            path=path,
            kind="xlsx",
            raw_text=raw_text,
            rendered_html=render_markdown(raw_text),
        )


def _worksheet_to_markdown(worksheet: Any) -> str:
    title = f"## {_escape_heading_text(worksheet.title)}"
    rows = list(_used_rows(worksheet))
    if not rows:
        return f"{title}\n\n빈 시트"

    column_count = max(len(row) for row in rows)
    normalized_rows = [_normalize_row(row, column_count) for row in rows]
    separator = ["---"] * column_count
    lines = [
        title,
        "",
        _format_table_row(normalized_rows[0]),
        _format_table_row(separator),
    ]
    lines.extend(_format_table_row(row) for row in normalized_rows[1:])
    return "\n".join(lines)


def _used_rows(worksheet: Any) -> Iterator[list[str]]:
    bounds = _used_bounds(worksheet)
    if bounds is None:
        return

    min_row, min_col, max_row, max_col = bounds
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        yield [_escape_table_cell(_cell_to_text(value)) for value in row]


def _used_bounds(worksheet: Any) -> tuple[int, int, int, int] | None:
    min_row: int | None = None
    min_col: int | None = None
    max_row = 0
    max_col = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if _is_empty_cell(cell.value):
                continue

            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)

    if min_row is None or min_col is None:
        return None
    return min_row, min_col, max_row, max_col


def _is_empty_cell(value: Any) -> bool:
    return value is None or value == ""


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _normalize_row(row: list[str], column_count: int) -> list[str]:
    return row + [""] * (column_count - len(row))


def _escape_heading_text(text: str) -> str:
    return text.replace("\r", " ").replace("\n", " ").strip()


def _escape_table_cell(text: str) -> str:
    return (
        text.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r\n", " ")
        .replace("\n", " ")
        .replace("\r", " ")
    )


def _format_table_row(row: list[str]) -> str:
    return "| " + " | ".join(row) + " |"
